from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import asyncio
import aiofiles
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
import uuid
import os
import google.generativeai as genai
import tempfile
import shutil
from datetime import datetime, timedelta
from dotenv import load_dotenv
from oletools.olevba import VBA_Parser
import re
import traceback
import hashlib
import threading
import time

# Charger les variables d'environnement
load_dotenv()

# Configuration
UPLOAD_MAX_SIZE = 50 * 1024 * 1024  # 50MB
ALLOWED_EXTENSIONS = {'.xlsx', '.xlsm'}
SESSION_TIMEOUT = 36000  # 10 heure (36000 secondes)
CLEANUP_INTERVAL = 300  # Nettoyage toutes les 5 minutes
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

# Initialize Gemini
if not GEMINI_API_KEY:
    raise Exception("❌ ERREUR CRITIQUE : Clé API Gemini manquante ! Ajoutez GEMINI_API_KEY dans les variables d'environnement")

genai.configure(api_key=GEMINI_API_KEY)

# Utiliser le bon modèle Gemini
model = None
models_to_try = ['gemini-1.5-flash', 'gemini-1.5-pro', 'gemini-pro']

for model_name in models_to_try:
    try:
        model = genai.GenerativeModel(model_name)
        test_response = model.generate_content("Test")
        print(f"✅ Modèle {model_name} initialisé avec succès")
        break
    except Exception as e:
        print(f"⚠️ Le modèle {model_name} n'est pas disponible : {str(e)[:100]}")
        continue

if not model:
    raise Exception("❌ Impossible d'initialiser un modèle Gemini. Vérifiez votre clé API.")

# Stockage en mémoire avec timestamps pour le nettoyage
sessions_store = {}
session_last_activity = {}

# ===== SYSTÈME DE NETTOYAGE AUTOMATIQUE =====

def cleanup_expired_sessions():
    """Nettoie les sessions expirées et leurs fichiers"""
    while True:
        try:
            current_time = datetime.now()
            expired_sessions = []
            
            for session_id, last_activity in session_last_activity.items():
                if current_time - last_activity > timedelta(seconds=SESSION_TIMEOUT):
                    expired_sessions.append(session_id)
            
            for session_id in expired_sessions:
                # Supprimer le fichier temporaire
                if session_id in sessions_store:
                    session_data = sessions_store[session_id]
                    file_path = session_data.get('file_path')
                    if file_path and os.path.exists(file_path):
                        try:
                            # Supprimer le fichier
                            os.remove(file_path)
                            # Supprimer le dossier temporaire s'il est vide
                            temp_dir = os.path.dirname(file_path)
                            if os.path.exists(temp_dir) and not os.listdir(temp_dir):
                                os.rmdir(temp_dir)
                            print(f"🗑️ Fichier nettoyé : {file_path}")
                        except Exception as e:
                            print(f"⚠️ Erreur lors de la suppression de {file_path}: {e}")
                
                # Supprimer de la mémoire
                sessions_store.pop(session_id, None)
                session_last_activity.pop(session_id, None)
                print(f"🗑️ Session expirée nettoyée : {session_id}")
            
            if expired_sessions:
                print(f"🧹 Nettoyage terminé : {len(expired_sessions)} sessions supprimées")
            
        except Exception as e:
            print(f"❌ Erreur lors du nettoyage : {e}")
        
        # Attendre avant le prochain nettoyage
        time.sleep(CLEANUP_INTERVAL)

# Démarrer le thread de nettoyage
cleanup_thread = threading.Thread(target=cleanup_expired_sessions, daemon=True)
cleanup_thread.start()
print(f"🧹 Système de nettoyage automatique démarré (timeout: {SESSION_TIMEOUT}s)")

# ===== NOUVELLES FONCTIONS D'ANALYSE INTELLIGENTE =====

def detect_content_type(headers: List[str], sample_data: List[List[str]]) -> str:
    """Détecte le type de contenu basé sur les en-têtes et données"""
    all_text = ' '.join(headers + [cell for row in sample_data for cell in row]).lower()
    
    if any(word in all_text for word in ['salaire', 'euro', 'montant', 'budget', 'dépense', 'coût', 'prix', 'taxes', 'logement', 'revenus']):
        return "📊 Données financières/budgétaires"
    elif any(word in all_text for word in ['client', 'nom', 'adresse', 'téléphone', 'email']):
        return "👥 Base de données clients"
    elif any(word in all_text for word in ['date', 'mois', 'planning', 'horaire', 'janvier', 'février', 'mars', 'avril', 'mai', 'juin']):
        return "📅 Planning/temporel"
    elif any(word in all_text for word in ['stock', 'quantité', 'produit', 'article', 'inventaire']):
        return "📦 Gestion de stock/inventaire"
    else:
        return "📋 Données mixtes"

def analyze_sheet_content(sheet: Dict[str, Any]) -> Dict[str, Any]:
    """Analyse le contenu d'une feuille et retourne un résumé détaillé"""
    data = sheet.get('data', [])
    if not data or len(data) < 1:
        return {
            'type': 'Feuille vide',
            'headers': [],
            'sample_data': [],
            'stats': 'Aucune donnée'
        }
    
    headers = data[0] if data else []
    sample_data = data[1:4] if len(data) > 1 else []
    content_type = detect_content_type(headers, sample_data)
    
    non_empty_cells = sum(1 for row in data for cell in row if cell and str(cell).strip())
    total_cells = sum(len(row) for row in data)
    fill_rate = (non_empty_cells / total_cells * 100) if total_cells > 0 else 0
    
    return {
        'type': content_type,
        'headers': headers[:6],
        'sample_data': sample_data,
        'stats': f"{len(data)} lignes, {len(headers)} colonnes, {fill_rate:.1f}% rempli",
        'formulas_count': len(sheet.get('formulas', {})),
        'has_formatting': len(sheet.get('formatting', {})) > 0
    }

def build_smart_context(session_data: Dict[str, Any]) -> str:
    """Construit un contexte intelligent basé sur les données réelles"""
    structure = session_data['structure']
    filename = session_data['filename']
    
    sheets_analysis = []
    for i, sheet in enumerate(structure['sheets']):
        analysis = analyze_sheet_content(sheet)
        
        sheet_summary = f"""
🔹 Feuille "{sheet['name']}" ({analysis['stats']})
   Type détecté : {analysis['type']}
   Colonnes : {', '.join(str(h) for h in analysis['headers'][:4])}{'...' if len(analysis['headers']) > 4 else ''}
   Formules : {analysis['formulas_count']}"""
        
        if i == 0 and analysis['sample_data']:
            sample_lines = []
            for row in analysis['sample_data'][:2]:
                row_preview = ' | '.join(str(cell)[:15] for cell in row[:4])
                sample_lines.append(f"   📄 {row_preview}")
            
            if sample_lines:
                sheet_summary += f"\n   Aperçu données :\n" + '\n'.join(sample_lines)
        
        sheets_analysis.append(sheet_summary)
    
    vba_info = ""
    vba_code = session_data.get('vba_code', {})
    
    if session_data.get('vba_modules') and vba_code:
        vba_modules = list(session_data['vba_modules'])
        vba_info = f"\n🔧 VBA : {len(vba_modules)} modules trouvés"
        
        for module_name in vba_modules[:3]:
            if module_name in vba_code:
                code = vba_code[module_name]
                code_lines = [line.strip() for line in code.split('\n') if line.strip() and not line.strip().startswith("'")]
                
                if code_lines:
                    code_preview = '\n'.join(code_lines[:5])
                    vba_info += f"""
   📝 Module "{module_name}" :
   {code_preview}{'...' if len(code_lines) > 5 else ''}"""
        
        if len(vba_modules) > 3:
            vba_info += f"\n   ... et {len(vba_modules) - 3} autres modules"
    elif session_data.get('vba_modules'):
        vba_modules = list(session_data['vba_modules'])
        vba_info = f"\n🔧 VBA : {len(vba_modules)} modules ({', '.join(vba_modules[:3])}{'...' if len(vba_modules) > 3 else ''})"
    
    context = f"""📊 FICHIER ANALYSÉ : {filename}
{''.join(sheets_analysis)}{vba_info}

🧠 CAPACITÉS ACTUELLES :
✅ Tu VOIS toutes ces données - ne demande JAMAIS de les décrire !
✅ Tu VOIS le code VBA ci-dessus - ne dis JAMAIS "sans connaître le code" !
✅ Tu peux analyser, calculer, optimiser directement
✅ Tu peux modifier les cellules via instructions naturelles
✅ Tu peux examiner et améliorer les formules/VBA existants"""

    return context

# ===== FONCTIONS UTILITAIRES =====

def get_color_hex(color_obj) -> Optional[str]:
    """Extrait la couleur hex d'un objet couleur openpyxl de manière sûre"""
    if not color_obj:
        return None
    
    try:
        if hasattr(color_obj, 'rgb') and isinstance(color_obj.rgb, str):
            rgb = color_obj.rgb
            if rgb == '00000000' or rgb == 'FF000000':
                return None
            if len(rgb) == 8:
                return f"#{rgb[2:]}"
            elif len(rgb) == 6:
                return f"#{rgb}"
        
        if hasattr(color_obj, 'value'):
            value = str(color_obj.value)
            if value and value != '00000000' and value != 'FF000000':
                if len(value) == 8:
                    return f"#{value[2:]}"
                elif len(value) == 6:
                    return f"#{value}"
        
        if hasattr(color_obj, 'r') and hasattr(color_obj, 'g') and hasattr(color_obj, 'b'):
            r = getattr(color_obj, 'r', 0)
            g = getattr(color_obj, 'g', 0)
            b = getattr(color_obj, 'b', 0)
            return f"#{r:02X}{g:02X}{b:02X}"
            
    except Exception as e:
        print(f"Erreur lors de l'extraction de la couleur: {e}")
        return None
    
    return None

def update_session_activity(session_id: str):
    """Met à jour l'activité d'une session"""
    session_last_activity[session_id] = datetime.now()

# ===== FASTAPI APP =====

app = FastAPI(
    title="Excel VBA Assistant API",
    description="API pour l'analyse et modification de fichiers Excel avec IA",
    version="1.0.0"
)

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Models
class ChatMessage(BaseModel):
    session_id: str
    message: str
    context: Optional[Dict[str, Any]] = None

class AnalyzeRequest(BaseModel):
    session_id: str
    module_name: Optional[str] = None
    sheet_name: Optional[str] = None

class CellUpdate(BaseModel):
    session_id: str
    sheet_name: str
    row: int
    col: int
    value: str

# Helper functions
async def get_session(session_id: str):
    update_session_activity(session_id)
    return sessions_store.get(session_id)

async def save_session(session_id: str, data: dict):
    sessions_store[session_id] = data
    update_session_activity(session_id)

def get_column_letter(col_idx: int) -> str:
    """Convert column index to Excel letter"""
    result = ""
    while col_idx >= 0:
        result = chr(col_idx % 26 + 65) + result
        col_idx = col_idx // 26 - 1
    return result

def extract_vba_code(workbook_path: str) -> Dict[str, str]:
    """Extract VBA code from Excel file using oletools"""
    vba_modules = {}
    try:
        vba_parser = VBA_Parser(workbook_path)
        if vba_parser.detect_vba_macros():
            for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                vba_modules[vba_filename] = vba_code
        vba_parser.close()
    except Exception as e:
        print(f"Error extracting VBA: {e}")
    return vba_modules

def analyze_excel_structure(workbook_path: str) -> Dict[str, Any]:
    """Analyze Excel file structure with full data and formatting"""
    wb = load_workbook(workbook_path, data_only=False, keep_vba=True)
    
    structure = {
        "sheets": [],
        "total_sheets": len(wb.sheetnames),
        "has_vba": workbook_path.endswith('.xlsm'),
        "file_size": os.path.getsize(workbook_path),
        "created": datetime.now().isoformat()
    }
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "has_data": sheet.max_row > 1 or sheet.max_column > 1,
            "formulas": {},
            "formatting": {},
            "charts": [],
            "tables": [],
            "data": [],
            "column_widths": {},
            "row_heights": {},
            "merged_cells": []
        }
        
        # Get column widths
        for col in sheet.column_dimensions:
            if sheet.column_dimensions[col].width:
                sheet_info["column_widths"][col] = sheet.column_dimensions[col].width
        
        # Get row heights
        for row in sheet.row_dimensions:
            if sheet.row_dimensions[row].height:
                sheet_info["row_heights"][row] = sheet.row_dimensions[row].height
        
        # Get merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_info["merged_cells"].append(str(merged_range))
        
        # Get all data
        sheet_data = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            sheet_data.append([str(cell) if cell is not None else "" for cell in row])
        
        sheet_info["data"] = sheet_data
        
        # Get formulas and formatting
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell_ref = cell.coordinate
                
                if hasattr(cell, 'data_type') and cell.data_type == 'f':
                    sheet_info["formulas"][cell_ref] = cell.value
                
                try:
                    cell_format = {}
                    
                    if cell.font:
                        font_info = {}
                        if cell.font.bold:
                            font_info['bold'] = True
                        if cell.font.italic:
                            font_info['italic'] = True
                        color = get_color_hex(cell.font.color)
                        if color:
                            font_info['color'] = color
                        if cell.font.size:
                            font_info['size'] = cell.font.size
                        if font_info:
                            cell_format['font'] = font_info
                    
                    if cell.fill and hasattr(cell.fill, 'fgColor'):
                        bg_color = get_color_hex(cell.fill.fgColor)
                        if bg_color:
                            cell_format['backgroundColor'] = bg_color
                    
                    if cell.alignment:
                        align_info = {}
                        if cell.alignment.horizontal:
                            align_info['horizontal'] = cell.alignment.horizontal
                        if cell.alignment.vertical:
                            align_info['vertical'] = cell.alignment.vertical
                        if align_info:
                            cell_format['alignment'] = align_info
                    
                    if cell.border:
                        border_info = {}
                        for side in ['left', 'right', 'top', 'bottom']:
                            border_side = getattr(cell.border, side)
                            if border_side and border_side.style:
                                side_info = {'style': border_side.style}
                                border_color = get_color_hex(border_side.color)
                                if border_color:
                                    side_info['color'] = border_color
                                border_info[side] = side_info
                        if border_info:
                            cell_format['border'] = border_info
                    
                    if cell.number_format and cell.number_format != 'General':
                        cell_format['numberFormat'] = cell.number_format
                    
                    if cell_format:
                        sheet_info["formatting"][cell_ref] = cell_format
                        
                except Exception as e:
                    print(f"Erreur lors du formatage de la cellule {cell_ref}: {str(e)}")
                    continue
        
        structure["sheets"].append(sheet_info)
    
    wb.close()
    return structure

async def generate_initial_analysis(filename: str, structure: Dict[str, Any], vba_modules: Dict[str, str]) -> str:
    """Generate initial analysis using smart context"""
    
    session_data = {
        'filename': filename,
        'structure': structure,
        'vba_modules': list(vba_modules.keys()) if vba_modules else []
    }
    
    smart_context = build_smart_context(session_data)
    
    prompt = f"""{smart_context}

🎯 MISSION : Fais une analyse EXPERTE et PRÉCISE de ce fichier Excel.

INSTRUCTIONS :
- Utilise les VRAIES données que tu vois
- Identifie le PURPOSE exact du fichier
- Trouve les problèmes/opportunités CONCRETS
- Propose 2-3 améliorations SPÉCIFIQUES
- Sois expert mais accessible (200-250 mots max)
- Utilise des émojis pour la clarté

ANALYSE ATTENDUE :
1. Purpose du fichier (basé sur les vraies données)
2. Points forts identifiés
3. Problèmes détectés (s'il y en a)
4. Suggestions d'amélioration concrètes

Fais une analyse digne d'un expert Excel ! 🎯"""
    
    response = await model.generate_content_async(prompt)
    return response.text

# ===== API ENDPOINTS =====

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload and analyze Excel file"""
    try:
        # Validate file
        if not any(file.filename.endswith(ext) for ext in ALLOWED_EXTENSIONS):
            raise HTTPException(400, "Invalid file type. Only .xlsx and .xlsm files are allowed.")
        
        # Check file size
        contents = await file.read()
        if len(contents) > UPLOAD_MAX_SIZE:
            raise HTTPException(400, f"File too large. Maximum size is {UPLOAD_MAX_SIZE // 1024 // 1024}MB")
        
        # Generate session ID
        session_id = str(uuid.uuid4())
        
        # Save file temporarily
        temp_dir = tempfile.mkdtemp()
        file_path = os.path.join(temp_dir, file.filename)
        
        async with aiofiles.open(file_path, 'wb') as f:
            await f.write(contents)
        
        # Analyze file structure
        try:
            structure = analyze_excel_structure(file_path)
        except Exception as e:
            print(f"Erreur lors de l'analyse du fichier Excel: {str(e)}")
            raise HTTPException(500, f"Erreur lors de l'analyse du fichier: {str(e)}")
        
        vba_modules = extract_vba_code(file_path) if file.filename.endswith('.xlsm') else {}
        
        # Generate initial analysis
        try:
            initial_analysis = await generate_initial_analysis(file.filename, structure, vba_modules)
        except Exception as e:
            print(f"Erreur lors de la génération de l'analyse initiale: {str(e)}")
            initial_analysis = f"📊 Fichier {file.filename} chargé avec succès ! Le fichier contient {len(structure['sheets'])} feuilles avec des données structurées. Utilisez le chat pour poser vos questions spécifiques."
        
        # Store session data
        session_data = {
            "filename": file.filename,
            "file_path": file_path,
            "structure": structure,
            "vba_modules": list(vba_modules.keys()),
            "vba_code": vba_modules,
            "created": datetime.now().isoformat(),
            "chat_history": []
        }
        
        await save_session(session_id, session_data)
        
        print(f"✅ Session créée : {session_id} - {file.filename}")
        
        return {
            "session_id": session_id,
            "filename": file.filename,
            "structure": structure,
            "vba_modules": list(vba_modules.keys()),
            "vba_code": vba_modules,
            "initial_analysis": initial_analysis
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Erreur inattendue dans upload_file: {str(e)}")
        traceback.print_exc()
        raise HTTPException(500, f"Erreur inattendue: {str(e)}")

@app.post("/api/chat")
async def chat_with_agent(message: ChatMessage):
    """Chat endpoint for conversational interactions"""
    session_data = await get_session(message.session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    # Analyser le message pour détecter une demande de modification
    message_lower = message.message.lower()
    modification_keywords = [
        'écris', 'ecris', 'écrire', 'ecrire',
        'mets', 'mettre', 'met',
        'change', 'changer', 'modifie', 'modifier',
        'insère', 'insere', 'insérer', 'inserer',
        'ajoute', 'ajouter',
        'remplace', 'remplacer',
        'saisis', 'saisir',
        'entre', 'entrer',
        'place', 'placer',
        'définis', 'definir', 'défini'
    ]
    
    has_cell_reference = bool(re.search(r'\b[A-Z]+\d+\b', message.message))
    is_modification = any(word in message_lower for word in modification_keywords) and has_cell_reference
    
    if is_modification:
        system_prompt = f"""
Tu es un assistant Excel. L'utilisateur veut modifier une cellule.
Réponds UNIQUEMENT avec ce format JSON, rien d'autre :

{{
    "action": "update_cell",
    "sheet": "Feuil1",
    "cell": "[LA_CELLULE]",
    "value": "[LA_VALEUR]",
    "message": "✅ J'ai modifié la cellule [LA_CELLULE] avec la valeur '[LA_VALEUR]'. La modification est sauvegardée."
}}

Message de l'utilisateur : {message.message}

IMPORTANT : Réponds UNIQUEMENT avec le JSON, pas de texte avant ou après !
"""
    else:
        smart_context = build_smart_context(session_data)
        
        system_prompt = f"""{smart_context}

🎯 COMPORTEMENT REQUIS :
- Utilise DIRECTEMENT les données que tu vois ci-dessus
- Ne demande JAMAIS "pouvez-vous me dire le contenu" - tu le vois déjà !
- Analyse et réponds de manière experte et proactive
- Propose des améliorations concrètes basées sur les VRAIES données
- Sois précis et utile, pas verbeux

❌ INTERDIT :
- Demander des informations sur le contenu (tu les as !)
- Dire "je n'ai pas accès" (tu as TOUT l'accès !)
- Poser des questions génériques sur le type de données

Question de l'utilisateur : "{message.message}"

Réponds en expert qui CONNAÎT parfaitement ce fichier ! 🚀"""
    
    # Generate response with Gemini
    response = await model.generate_content_async(system_prompt)
    response_text = response.text.strip()
    
    # Vérifier si la réponse contient une action
    try:
        json_match = re.search(r'\{[^{}]*"action"[^{}]*\}', response_text, re.DOTALL)
        if json_match:
            action_data = json.loads(json_match.group())
            
            if action_data.get('action') == 'update_cell':
                # Effectuer la modification
                sheet_name = action_data.get('sheet', 'Feuil1')
                cell_ref = action_data.get('cell', '').upper()
                value = action_data.get('value', '')
                
                # Convertir la référence de cellule en indices
                col_letters = ''.join(filter(str.isalpha, cell_ref))
                row_num = int(''.join(filter(str.isdigit, cell_ref))) - 1
                
                col_num = 0
                for char in col_letters:
                    col_num = col_num * 26 + (ord(char.upper()) - ord('A') + 1)
                col_num -= 1
                
                # Mettre à jour la cellule
                sheet_found = False
                for sheet_data in session_data['structure']['sheets']:
                    if sheet_data['name'] == sheet_name:
                        sheet_found = True
                        while len(sheet_data['data']) <= row_num:
                            sheet_data['data'].append([''] * (col_num + 1))
                        
                        while len(sheet_data['data'][row_num]) <= col_num:
                            sheet_data['data'][row_num].append('')
                        
                        sheet_data['data'][row_num][col_num] = str(value)
                        sheet_data['max_row'] = max(sheet_data['max_row'], row_num + 1)
                        sheet_data['max_column'] = max(sheet_data['max_column'], col_num + 1)
                        
                        if value.startswith('='):
                            sheet_data['formulas'][cell_ref] = value
                        elif cell_ref in sheet_data['formulas']:
                            del sheet_data['formulas'][cell_ref]
                        
                        break
                
                if sheet_found:
                    # Mettre à jour le fichier Excel
                    wb = load_workbook(session_data['file_path'], keep_vba=True)
                    ws = wb[sheet_name]
                    
                    if value.startswith('='):
                        ws[cell_ref] = value
                    else:
                        try:
                            if '.' in value:
                                ws[cell_ref] = float(value)
                            else:
                                ws[cell_ref] = int(value)
                        except ValueError:
                            ws[cell_ref] = value
                    
                    wb.save(session_data['file_path'])
                    wb.close()
                    
                    await save_session(message.session_id, session_data)
                    response_text = action_data.get('message', f"✅ J'ai modifié la cellule {cell_ref} avec la valeur '{value}'")
                else:
                    response_text = f"❌ Erreur : La feuille {sheet_name} n'existe pas."
    except Exception as e:
        pass
    
    # Update chat history
    session_data['chat_history'].append({
        "user": message.message,
        "assistant": response_text,
        "timestamp": datetime.now().isoformat()
    })
    
    await save_session(message.session_id, session_data)
    
    # Stream response
    async def generate():
        for chunk in response_text.split(' '):
            yield f"data: {json.dumps({'chunk': chunk + ' '})}\n\n"
            await asyncio.sleep(0.01)
        yield f"data: {json.dumps({'done': True})}\n\n"
    
    return StreamingResponse(generate(), media_type="text/event-stream")

@app.post("/api/update-cell")
async def update_cell(update: CellUpdate):
    """Update a single cell value"""
    session_data = await get_session(update.session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    try:
        # Update in memory
        for sheet in session_data['structure']['sheets']:
            if sheet['name'] == update.sheet_name:
                while len(sheet['data']) <= update.row:
                    sheet['data'].append([''] * (update.col + 1))
                
                while len(sheet['data'][update.row]) <= update.col:
                    sheet['data'][update.row].append('')
                
                sheet['data'][update.row][update.col] = update.value
                sheet['max_row'] = max(sheet['max_row'], update.row + 1)
                sheet['max_column'] = max(sheet['max_column'], update.col + 1)
                
                cell_ref = f"{get_column_letter(update.col)}{update.row + 1}"
                if update.value.startswith('='):
                    sheet['formulas'][cell_ref] = update.value
                elif cell_ref in sheet['formulas']:
                    del sheet['formulas'][cell_ref]
                break
        
        # Update the actual Excel file
        wb = load_workbook(session_data['file_path'], keep_vba=True)
        ws = wb[update.sheet_name]
        
        cell = ws.cell(row=update.row + 1, column=update.col + 1)
        
        if update.value.startswith('='):
            cell.value = update.value
            cell.data_type = 'f'
        else:
            try:
                if '.' in update.value:
                    cell.value = float(update.value)
                else:
                    cell.value = int(update.value)
            except ValueError:
                cell.value = update.value
        
        wb.save(session_data['file_path'])
        wb.close()
        
        await save_session(update.session_id, session_data)
        
        return {"status": "success", "message": "Cell updated"}
    except Exception as e:
        raise HTTPException(500, f"Error updating cell: {str(e)}")

@app.get("/api/sheet-data/{session_id}/{sheet_name}")
async def get_sheet_data(
    session_id: str,
    sheet_name: str,
    start_row: int = 0,
    end_row: Optional[int] = None,
    start_col: int = 0,
    end_col: Optional[int] = None
):
    """Get sheet data with pagination support"""
    session_data = await get_session(session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    sheet = None
    for s in session_data['structure']['sheets']:
        if s['name'] == sheet_name:
            sheet = s
            break
    
    if not sheet:
        raise HTTPException(404, f"Sheet {sheet_name} not found")
    
    data = sheet['data']
    
    if end_row is None:
        end_row = len(data)
    if end_col is None:
        end_col = max(len(row) for row in data) if data else 0
    
    sliced_data = []
    for row_idx in range(start_row, min(end_row, len(data))):
        row = data[row_idx]
        sliced_row = row[start_col:min(end_col, len(row))]
        if len(sliced_row) < (end_col - start_col):
            sliced_row.extend([''] * ((end_col - start_col) - len(sliced_row)))
        sliced_data.append(sliced_row)
    
    return {
        "data": sliced_data,
        "total_rows": len(data),
        "total_cols": max(len(row) for row in data) if data else 0,
        "formulas": sheet.get('formulas', {}),
        "formatting": sheet.get('formatting', {}),
        "column_widths": sheet.get('column_widths', {}),
        "row_heights": sheet.get('row_heights', {}),
        "merged_cells": sheet.get('merged_cells', [])
    }

@app.post("/api/recalculate/{session_id}")
async def recalculate_formulas(session_id: str):
    """Recalculate all formulas in the workbook"""
    session_data = await get_session(session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    try:
        wb = load_workbook(session_data['file_path'], data_only=False, keep_vba=True)
        wb.calculation.calcMode = 'automatic'
        wb.save(session_data['file_path'])
        wb.close()
        
        wb = load_workbook(session_data['file_path'], data_only=True, keep_vba=True)
        
        for sheet_data in session_data['structure']['sheets']:
            sheet_name = sheet_data['name']
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                new_data = []
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                    new_data.append([str(cell) if cell is not None else "" for cell in row])
                sheet_data['data'] = new_data
        
        wb.close()
        await save_session(session_id, session_data)
        
        return {"status": "success", "message": "Formulas recalculated"}
    except Exception as e:
        raise HTTPException(500, f"Error recalculating formulas: {str(e)}")

@app.post("/api/export")
async def export_file(session_id: str):
    """Export modified Excel file"""
    session_data = await get_session(session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    file_path = session_data['file_path']
    
    if not os.path.exists(file_path):
        raise HTTPException(404, "File not found")
    
    async def iterfile():
        async with aiofiles.open(file_path, 'rb') as f:
            while chunk := await f.read(8192):
                yield chunk
    
    return StreamingResponse(
        iterfile(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=modified_{session_data['filename']}"
        }
    )

@app.get("/api/session/{session_id}")
async def get_session_data(session_id: str):
    """Get session data"""
    session_data = await get_session(session_id)
    
    if not session_data:
        raise HTTPException(404, "Session not found or expired")
    
    return session_data

@app.get("/api/stats")
async def get_stats():
    """Get API statistics"""
    return {
        "active_sessions": len(sessions_store),
        "session_timeout": SESSION_TIMEOUT,
        "cleanup_interval": CLEANUP_INTERVAL,
        "max_file_size": f"{UPLOAD_MAX_SIZE // 1024 // 1024}MB"
    }

@app.get("/")
async def root():
    return {
        "message": "Excel VBA Assistant API is running!",
        "gemini_status": "✅ Gemini AI activé" if model else "❌ Gemini AI non disponible",
        "session_management": "✅ Nettoyage automatique activé",
        "active_sessions": len(sessions_store),
        "features": [
            "✅ Support complet des données Excel",
            "✅ Formules Excel",
            "✅ Formatage des cellules",
            "✅ Chargement par chunks",
            "✅ Modification via chat IA",
            "✅ Analyse contextuelle intelligente",
            "✅ Nettoyage automatique des sessions"
        ]
    }

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)